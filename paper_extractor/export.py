"""
paper_extractor/export.py -- Export database to Excel
======================================================
"""

import sqlite3
import json
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

from paper_extractor.db import get_db_path, init_db


# Header style
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
)


def export_excel(db_dir: str, output_path: Optional[str] = None, verbose: bool = True):
    """
    Export all database tables to an Excel workbook.

    Sheets:
      1. Summary - counts and stats
      2. Binding Data - all binding thermodynamics
      3. Physical Constants - dissolution, solvation, barriers
      4. Papers - processed paper list
      5. Raw Tables - unclassified/low-confidence tables for review
      6. Review Queue - items needing manual attention
    """
    if Workbook is None:
        raise ImportError("Need openpyxl. Install: pip install openpyxl")

    db_path = get_db_path(db_dir)
    conn = init_db(db_path)

    if output_path is None:
        output_path = str(Path(db_dir) / "extracted_data.xlsx")

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    _write_summary(ws, conn)

    # ── Sheet 2: Binding Data ──
    ws2 = wb.create_sheet("Binding Data")
    _write_query_sheet(ws2, conn, """
        SELECT b.bind_id, p.filename, p.doi, p.year,
               b.receptor, b.ligand, b.Ka, b.Kd, b.Kd_unit,
               b.dG, b.dG_unit, b.dH, b.dH_unit, b.TdS, b.TdS_unit,
               b.dCp, b.n_sites, b.temperature_C, b.pH,
               b.method, b.table_ref, b.confidence, b.raw_text
        FROM binding_data b
        JOIN papers p ON b.paper_id = p.paper_id
        ORDER BY b.confidence DESC, p.year, b.receptor, b.ligand
    """, headers=[
        'ID', 'File', 'DOI', 'Year', 'Receptor', 'Ligand',
        'Ka (M-1)', 'Kd', 'Kd unit', 'dG', 'dG unit', 'dH', 'dH unit',
        'TdS', 'TdS unit', 'dCp', 'n_sites', 'Temp (C)', 'pH',
        'Method', 'Table Ref', 'Confidence', 'Raw Text',
    ])

    # ── Sheet 3: Physical Constants ──
    ws3 = wb.create_sheet("Physical Constants")
    _write_query_sheet(ws3, conn, """
        SELECT c.const_id, p.filename, p.doi, p.year,
               c.quantity, c.compound, c.value, c.unit,
               c.temperature_C, c.solvent, c.conditions,
               c.table_ref, c.confidence, c.raw_text
        FROM physical_constants c
        JOIN papers p ON c.paper_id = p.paper_id
        ORDER BY c.quantity, c.compound
    """, headers=[
        'ID', 'File', 'DOI', 'Year', 'Quantity', 'Compound',
        'Value', 'Unit', 'Temp (C)', 'Solvent', 'Conditions',
        'Table Ref', 'Confidence', 'Raw Text',
    ])

    # ── Sheet 4: Papers ──
    ws4 = wb.create_sheet("Papers")
    _write_query_sheet(ws4, conn, """
        SELECT paper_id, filename, title, authors, year, journal, doi,
               n_pages, status, first_processed
        FROM papers ORDER BY year, filename
    """, headers=[
        'ID', 'Filename', 'Title', 'Authors', 'Year', 'Journal', 'DOI',
        'Pages', 'Status', 'Processed',
    ])

    # ── Sheet 5: Raw Tables (for review) ──
    ws5 = wb.create_sheet("Raw Tables")
    rows = conn.execute("""
        SELECT r.table_id, p.filename, r.page_number, r.table_index,
               r.classified_as, r.confidence, r.extracted, r.headers, r.rows
        FROM raw_tables r
        JOIN papers p ON r.paper_id = p.paper_id
        WHERE r.extracted = 0
        ORDER BY r.classified_as, r.confidence DESC
    """).fetchall()

    raw_headers = ['ID', 'File', 'Page', 'Table#', 'Classification',
                   'Confidence', 'Extracted', 'Headers', 'First Row']
    _write_headers(ws5, raw_headers)
    for i, row in enumerate(rows, start=2):
        ws5.cell(row=i, column=1, value=row['table_id'])
        ws5.cell(row=i, column=2, value=row['filename'])
        ws5.cell(row=i, column=3, value=row['page_number'])
        ws5.cell(row=i, column=4, value=row['table_index'])
        ws5.cell(row=i, column=5, value=row['classified_as'])
        ws5.cell(row=i, column=6, value=row['confidence'])
        ws5.cell(row=i, column=7, value=row['extracted'])
        # Headers as readable string
        try:
            hdrs = json.loads(row['headers'])
            ws5.cell(row=i, column=8, value=' | '.join(str(h) for h in hdrs[:8]))
        except (json.JSONDecodeError, TypeError):
            ws5.cell(row=i, column=8, value=str(row['headers'])[:200])
        # First data row
        try:
            rws = json.loads(row['rows'])
            if rws:
                ws5.cell(row=i, column=9, value=' | '.join(str(c) for c in rws[0][:8]))
        except (json.JSONDecodeError, TypeError):
            pass

    _auto_width(ws5)

    # ── Sheet 6: Review Queue ──
    ws6 = wb.create_sheet("Review Queue")
    _write_query_sheet(ws6, conn, """
        SELECT rq.review_id, p.filename, rq.source_table, rq.source_id,
               rq.issue, rq.context, rq.resolved, rq.resolution
        FROM review_queue rq
        JOIN papers p ON rq.paper_id = p.paper_id
        WHERE rq.resolved = 0
        ORDER BY rq.review_id
    """, headers=[
        'ID', 'File', 'Source Table', 'Source ID',
        'Issue', 'Context', 'Resolved', 'Resolution',
    ])

    # Save
    wb.save(output_path)
    conn.close()

    if verbose:
        print(f"Exported to: {output_path}")

    return output_path


def _write_headers(ws, headers):
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_query_sheet(ws, conn, query, headers):
    _write_headers(ws, headers)
    rows = conn.execute(query).fetchall()
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    _auto_width(ws)


def _auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:50]:  # check first 50 rows only
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_summary(ws, conn):
    ws.cell(row=1, column=1, value="MABE Paper Data Extraction Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    from paper_extractor.db import get_stats
    stats = get_stats(conn)

    rows = [
        ("Papers processed", stats.get('papers', 0)),
        ("Binding data entries", stats.get('binding_data', 0)),
        ("Physical constants", stats.get('physical_constants', 0)),
        ("Crystal contacts", stats.get('contacts', 0)),
        ("Raw tables stored", stats.get('raw_tables', 0)),
        ("Items needing review", stats.get('unresolved_reviews', 0)),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=1).font = Font(bold=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
