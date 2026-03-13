#!/usr/bin/env python3
"""
bootstrap_paper_extractor.py -- Deploy paper extraction tool (v0.2)

Improvements in v0.2:
  - Text normalization fixes split tokens (K\na -> Ka)
  - Scientific notation handling (4.9 x 10^5 -> 490000)
  - Context-aware ligand/receptor name extraction
  - Garbled table parser for pipe-separated mangled tables
  - dH, dG, TdS extraction from running text
  - Dedup by (field, value) to avoid repeat entries
  - Sentence-level context windowing
"""
import os

FILES = {}

FILES["paper_extractor/__init__.py"] = r'''
"""
paper_extractor/__init__.py
============================
PDF data extraction tool for MABE physics parameters.
Reads papers from a folder, extracts structured data, stores in SQLite.
"""

__version__ = "0.1.0"
'''

FILES["paper_extractor/__main__.py"] = r'''
"""
paper_extractor/__main__.py -- CLI interface
=============================================

Usage:
  python -m paper_extractor run [folder]       Process PDFs in folder
  python -m paper_extractor export [folder]    Export to Excel
  python -m paper_extractor stats [folder]     Show database stats
  python -m paper_extractor review [folder]    Show items needing review

Default folder: ./papers/
"""

import sys
import os
from pathlib import Path


DEFAULT_FOLDER = "papers"


def main():
    args = sys.argv[1:]

    if not args or args[0] in ('-h', '--help', 'help'):
        print(__doc__)
        return

    command = args[0]
    folder = args[1] if len(args) > 1 else DEFAULT_FOLDER

    # Resolve to absolute path
    folder = str(Path(folder).resolve())

    if command == 'run':
        cmd_run(folder)
    elif command == 'export':
        cmd_export(folder)
    elif command == 'stats':
        cmd_stats(folder)
    elif command == 'review':
        cmd_review(folder)
    else:
        print(f"Unknown command: {command}")
        print(__doc__)
        sys.exit(1)


def cmd_run(folder: str):
    """Process all PDFs in folder."""
    if not os.path.isdir(folder):
        print(f"Folder not found: {folder}")
        print(f"Create it and put PDFs inside, then run again.")
        sys.exit(1)

    pdfs = list(Path(folder).glob("*.pdf"))
    print(f"Paper Extractor v0.1")
    print(f"Folder: {folder}")
    print(f"PDFs found: {len(pdfs)}")
    print()

    from paper_extractor.extractor import process_folder
    counts = process_folder(folder, verbose=True)

    print()
    print("Run 'python -m paper_extractor export' to generate Excel output.")
    print("Run 'python -m paper_extractor review' to see items needing attention.")


def cmd_export(folder: str):
    """Export database to Excel."""
    from paper_extractor.db import get_db_path
    db_path = get_db_path(folder)
    if not os.path.exists(db_path):
        print(f"No database found at {db_path}")
        print("Run 'python -m paper_extractor run' first.")
        sys.exit(1)

    from paper_extractor.export import export_excel
    output = export_excel(folder, verbose=True)
    print(f"\nOpen {output} to review extracted data.")


def cmd_stats(folder: str):
    """Show database statistics."""
    from paper_extractor.db import get_db_path, init_db, get_stats
    db_path = get_db_path(folder)
    if not os.path.exists(db_path):
        print(f"No database found at {db_path}")
        return

    conn = init_db(db_path)
    stats = get_stats(conn)
    conn.close()

    print(f"Paper Extractor Database: {db_path}")
    print(f"  Papers:            {stats['papers']}")
    print(f"  Binding entries:   {stats['binding_data']}")
    print(f"  Physical constants:{stats['physical_constants']}")
    print(f"  Crystal contacts:  {stats['contacts']}")
    print(f"  Raw tables:        {stats['raw_tables']}")
    print(f"  Needs review:      {stats['unresolved_reviews']}")


def cmd_review(folder: str):
    """Show items needing manual review."""
    from paper_extractor.db import get_db_path, init_db
    db_path = get_db_path(folder)
    if not os.path.exists(db_path):
        print(f"No database found at {db_path}")
        return

    conn = init_db(db_path)
    rows = conn.execute("""
        SELECT rq.review_id, p.filename, rq.issue, rq.context
        FROM review_queue rq
        JOIN papers p ON rq.paper_id = p.paper_id
        WHERE rq.resolved = 0
        ORDER BY rq.review_id
    """).fetchall()
    conn.close()

    if not rows:
        print("No items needing review.")
        return

    print(f"{len(rows)} items needing review:\n")
    for row in rows:
        print(f"  [{row['review_id']}] {row['filename']}")
        print(f"       Issue: {row['issue']}")
        if row['context']:
            print(f"       Context: {row['context'][:100]}")
        print()


if __name__ == '__main__':
    main()
'''

FILES["paper_extractor/db.py"] = r'''
"""
paper_extractor/db.py -- SQLite database for extracted data
============================================================
"""

import sqlite3
import hashlib
import os
import json
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any


DB_NAME = "paper_data.db"


def get_db_path(base_dir: str) -> str:
    return os.path.join(base_dir, DB_NAME)


def init_db(db_path: str) -> sqlite3.Connection:
    """Create database and tables if they don't exist."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    conn.executescript("""
    -- Papers we've processed
    CREATE TABLE IF NOT EXISTS papers (
        paper_id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT NOT NULL,
        file_hash TEXT UNIQUE NOT NULL,
        title TEXT,
        authors TEXT,
        year INTEGER,
        journal TEXT,
        doi TEXT,
        first_processed TEXT NOT NULL,
        last_processed TEXT NOT NULL,
        n_pages INTEGER,
        status TEXT DEFAULT 'processed'  -- 'processed', 'failed', 'needs_review'
    );

    -- Binding thermodynamics (ITC, SPR, fluorescence)
    CREATE TABLE IF NOT EXISTS binding_data (
        bind_id INTEGER PRIMARY KEY AUTOINCREMENT,
        paper_id INTEGER NOT NULL REFERENCES papers(paper_id),
        receptor TEXT,           -- lectin name, host name, protein
        ligand TEXT,             -- sugar, guest molecule
        Ka REAL,                 -- association constant M^-1
        Kd REAL,                 -- dissociation constant M or mM
        Kd_unit TEXT DEFAULT 'M',
        dG REAL,                 -- kcal/mol (as published)
        dG_unit TEXT DEFAULT 'kcal/mol',
        dH REAL,                 -- kcal/mol
        dH_unit TEXT DEFAULT 'kcal/mol',
        TdS REAL,                -- kcal/mol
        TdS_unit TEXT DEFAULT 'kcal/mol',
        dCp REAL,                -- cal/mol/K
        n_sites REAL,            -- stoichiometry
        temperature_C REAL,
        pH REAL,
        buffer TEXT,
        method TEXT,             -- 'ITC', 'SPR', 'fluorescence', 'NMR', 'inhibition'
        conditions TEXT,         -- freeform notes
        table_ref TEXT,          -- 'Table 1', 'Table 3', etc.
        confidence TEXT DEFAULT 'medium',  -- 'high', 'medium', 'low'
        raw_text TEXT,           -- original text snippet for verification
        UNIQUE(paper_id, receptor, ligand, temperature_C, method)
    );

    -- Physical constants (desolvation, solvation, transfer energies)
    CREATE TABLE IF NOT EXISTS physical_constants (
        const_id INTEGER PRIMARY KEY AUTOINCREMENT,
        paper_id INTEGER NOT NULL REFERENCES papers(paper_id),
        quantity TEXT NOT NULL,   -- 'dH_sol', 'dG_hyd', 'dCp', 'Ka', 'barrier_height'
        compound TEXT NOT NULL,
        value REAL NOT NULL,
        unit TEXT NOT NULL,
        temperature_C REAL,
        solvent TEXT DEFAULT 'water',
        conditions TEXT,
        table_ref TEXT,
        confidence TEXT DEFAULT 'medium',
        raw_text TEXT
    );

    -- Crystal structure contacts
    CREATE TABLE IF NOT EXISTS contacts (
        contact_id INTEGER PRIMARY KEY AUTOINCREMENT,
        paper_id INTEGER NOT NULL REFERENCES papers(paper_id),
        pdb_id TEXT,
        receptor TEXT,
        ligand TEXT,
        residue_position TEXT,   -- 'C3-OH', 'C4-OH', etc.
        contact_type TEXT,       -- 'hbond', 'ch_pi', 'vdw', 'water_bridge', 'metal'
        partner_residue TEXT,    -- 'Asn14', 'Trp62', 'Ca2+'
        distance_A REAL,
        angle_deg REAL,
        notes TEXT,
        confidence TEXT DEFAULT 'medium',
        raw_text TEXT
    );

    -- Raw table extractions (for manual review)
    CREATE TABLE IF NOT EXISTS raw_tables (
        table_id INTEGER PRIMARY KEY AUTOINCREMENT,
        paper_id INTEGER NOT NULL REFERENCES papers(paper_id),
        page_number INTEGER,
        table_index INTEGER,     -- nth table on that page
        headers TEXT,            -- JSON list of column headers
        rows TEXT,               -- JSON list of row lists
        classified_as TEXT,      -- 'binding', 'physical', 'contacts', 'other', 'unclassified'
        confidence TEXT DEFAULT 'low',
        extracted INTEGER DEFAULT 0  -- 1 if data was successfully parsed into typed tables
    );

    -- Manual review queue
    CREATE TABLE IF NOT EXISTS review_queue (
        review_id INTEGER PRIMARY KEY AUTOINCREMENT,
        paper_id INTEGER NOT NULL REFERENCES papers(paper_id),
        source_table TEXT,       -- which table the issue is from
        source_id INTEGER,
        issue TEXT,              -- 'ambiguous_unit', 'missing_value', 'unusual_format'
        context TEXT,            -- surrounding text
        resolved INTEGER DEFAULT 0,
        resolution TEXT
    );
    """)

    conn.commit()
    return conn


def file_hash(filepath: str) -> str:
    """SHA256 hash of file contents for dedup."""
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()


def is_paper_processed(conn: sqlite3.Connection, fhash: str) -> bool:
    """Check if a paper has already been processed."""
    row = conn.execute("SELECT paper_id FROM papers WHERE file_hash = ?", (fhash,)).fetchone()
    return row is not None


def register_paper(conn: sqlite3.Connection, filename: str, fhash: str,
                   n_pages: int, title: str = "", authors: str = "",
                   year: int = None, journal: str = "", doi: str = "") -> int:
    """Register a new paper. Returns paper_id."""
    now = datetime.now().isoformat()
    cur = conn.execute("""
        INSERT INTO papers (filename, file_hash, title, authors, year, journal, doi,
                           first_processed, last_processed, n_pages, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'processed')
    """, (filename, fhash, title, authors, year, journal, doi, now, now, n_pages))
    conn.commit()
    return cur.lastrowid


def insert_binding(conn: sqlite3.Connection, paper_id: int, data: Dict[str, Any]) -> Optional[int]:
    """Insert a binding data row. Returns bind_id or None on conflict."""
    cols = ['paper_id', 'receptor', 'ligand', 'Ka', 'Kd', 'Kd_unit',
            'dG', 'dG_unit', 'dH', 'dH_unit', 'TdS', 'TdS_unit',
            'dCp', 'n_sites', 'temperature_C', 'pH', 'buffer', 'method',
            'conditions', 'table_ref', 'confidence', 'raw_text']
    data['paper_id'] = paper_id
    vals = [data.get(c) for c in cols]
    placeholders = ','.join(['?'] * len(cols))
    colnames = ','.join(cols)
    try:
        cur = conn.execute(f"INSERT OR IGNORE INTO binding_data ({colnames}) VALUES ({placeholders})", vals)
        conn.commit()
        return cur.lastrowid if cur.rowcount > 0 else None
    except Exception as e:
        print(f"  Warning: binding insert failed: {e}")
        return None


def insert_physical_constant(conn: sqlite3.Connection, paper_id: int, data: Dict[str, Any]) -> Optional[int]:
    cols = ['paper_id', 'quantity', 'compound', 'value', 'unit',
            'temperature_C', 'solvent', 'conditions', 'table_ref',
            'confidence', 'raw_text']
    data['paper_id'] = paper_id
    vals = [data.get(c) for c in cols]
    placeholders = ','.join(['?'] * len(cols))
    colnames = ','.join(cols)
    try:
        cur = conn.execute(f"INSERT INTO physical_constants ({colnames}) VALUES ({placeholders})", vals)
        conn.commit()
        return cur.lastrowid
    except Exception as e:
        print(f"  Warning: constant insert failed: {e}")
        return None


def insert_raw_table(conn: sqlite3.Connection, paper_id: int,
                     page_number: int, table_index: int,
                     headers: list, rows: list,
                     classified_as: str = 'unclassified',
                     confidence: str = 'low') -> int:
    cur = conn.execute("""
        INSERT INTO raw_tables (paper_id, page_number, table_index,
                               headers, rows, classified_as, confidence)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (paper_id, page_number, table_index,
          json.dumps(headers), json.dumps(rows),
          classified_as, confidence))
    conn.commit()
    return cur.lastrowid


def add_review_item(conn: sqlite3.Connection, paper_id: int,
                    source_table: str, source_id: int,
                    issue: str, context: str = ""):
    conn.execute("""
        INSERT INTO review_queue (paper_id, source_table, source_id, issue, context)
        VALUES (?, ?, ?, ?, ?)
    """, (paper_id, source_table, source_id, issue, context))
    conn.commit()


def get_stats(conn: sqlite3.Connection) -> Dict[str, int]:
    """Get database statistics."""
    stats = {}
    for table in ['papers', 'binding_data', 'physical_constants', 'contacts',
                  'raw_tables', 'review_queue']:
        row = conn.execute(f"SELECT COUNT(*) as n FROM {table}").fetchone()
        stats[table] = row['n']
    # Unresolved reviews
    row = conn.execute("SELECT COUNT(*) as n FROM review_queue WHERE resolved=0").fetchone()
    stats['unresolved_reviews'] = row['n']
    return stats
'''

FILES["paper_extractor/export.py"] = r'''
"""
paper_extractor/export.py -- Export database to Excel
======================================================
"""

import sqlite3
import json
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

from paper_extractor.db import get_db_path, init_db


# Header style
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
)


def export_excel(db_dir: str, output_path: Optional[str] = None, verbose: bool = True):
    """
    Export all database tables to an Excel workbook.

    Sheets:
      1. Summary - counts and stats
      2. Binding Data - all binding thermodynamics
      3. Physical Constants - dissolution, solvation, barriers
      4. Papers - processed paper list
      5. Raw Tables - unclassified/low-confidence tables for review
      6. Review Queue - items needing manual attention
    """
    if Workbook is None:
        raise ImportError("Need openpyxl. Install: pip install openpyxl")

    db_path = get_db_path(db_dir)
    conn = init_db(db_path)

    if output_path is None:
        output_path = str(Path(db_dir) / "extracted_data.xlsx")

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    _write_summary(ws, conn)

    # ── Sheet 2: Binding Data ──
    ws2 = wb.create_sheet("Binding Data")
    _write_query_sheet(ws2, conn, """
        SELECT b.bind_id, p.filename, p.doi, p.year,
               b.receptor, b.ligand, b.Ka, b.Kd, b.Kd_unit,
               b.dG, b.dG_unit, b.dH, b.dH_unit, b.TdS, b.TdS_unit,
               b.dCp, b.n_sites, b.temperature_C, b.pH,
               b.method, b.table_ref, b.confidence, b.raw_text
        FROM binding_data b
        JOIN papers p ON b.paper_id = p.paper_id
        ORDER BY b.confidence DESC, p.year, b.receptor, b.ligand
    """, headers=[
        'ID', 'File', 'DOI', 'Year', 'Receptor', 'Ligand',
        'Ka (M-1)', 'Kd', 'Kd unit', 'dG', 'dG unit', 'dH', 'dH unit',
        'TdS', 'TdS unit', 'dCp', 'n_sites', 'Temp (C)', 'pH',
        'Method', 'Table Ref', 'Confidence', 'Raw Text',
    ])

    # ── Sheet 3: Physical Constants ──
    ws3 = wb.create_sheet("Physical Constants")
    _write_query_sheet(ws3, conn, """
        SELECT c.const_id, p.filename, p.doi, p.year,
               c.quantity, c.compound, c.value, c.unit,
               c.temperature_C, c.solvent, c.conditions,
               c.table_ref, c.confidence, c.raw_text
        FROM physical_constants c
        JOIN papers p ON c.paper_id = p.paper_id
        ORDER BY c.quantity, c.compound
    """, headers=[
        'ID', 'File', 'DOI', 'Year', 'Quantity', 'Compound',
        'Value', 'Unit', 'Temp (C)', 'Solvent', 'Conditions',
        'Table Ref', 'Confidence', 'Raw Text',
    ])

    # ── Sheet 4: Papers ──
    ws4 = wb.create_sheet("Papers")
    _write_query_sheet(ws4, conn, """
        SELECT paper_id, filename, title, authors, year, journal, doi,
               n_pages, status, first_processed
        FROM papers ORDER BY year, filename
    """, headers=[
        'ID', 'Filename', 'Title', 'Authors', 'Year', 'Journal', 'DOI',
        'Pages', 'Status', 'Processed',
    ])

    # ── Sheet 5: Raw Tables (for review) ──
    ws5 = wb.create_sheet("Raw Tables")
    rows = conn.execute("""
        SELECT r.table_id, p.filename, r.page_number, r.table_index,
               r.classified_as, r.confidence, r.extracted, r.headers, r.rows
        FROM raw_tables r
        JOIN papers p ON r.paper_id = p.paper_id
        WHERE r.extracted = 0
        ORDER BY r.classified_as, r.confidence DESC
    """).fetchall()

    raw_headers = ['ID', 'File', 'Page', 'Table#', 'Classification',
                   'Confidence', 'Extracted', 'Headers', 'First Row']
    _write_headers(ws5, raw_headers)
    for i, row in enumerate(rows, start=2):
        ws5.cell(row=i, column=1, value=row['table_id'])
        ws5.cell(row=i, column=2, value=row['filename'])
        ws5.cell(row=i, column=3, value=row['page_number'])
        ws5.cell(row=i, column=4, value=row['table_index'])
        ws5.cell(row=i, column=5, value=row['classified_as'])
        ws5.cell(row=i, column=6, value=row['confidence'])
        ws5.cell(row=i, column=7, value=row['extracted'])
        # Headers as readable string
        try:
            hdrs = json.loads(row['headers'])
            ws5.cell(row=i, column=8, value=' | '.join(str(h) for h in hdrs[:8]))
        except (json.JSONDecodeError, TypeError):
            ws5.cell(row=i, column=8, value=str(row['headers'])[:200])
        # First data row
        try:
            rws = json.loads(row['rows'])
            if rws:
                ws5.cell(row=i, column=9, value=' | '.join(str(c) for c in rws[0][:8]))
        except (json.JSONDecodeError, TypeError):
            pass

    _auto_width(ws5)

    # ── Sheet 6: Review Queue ──
    ws6 = wb.create_sheet("Review Queue")
    _write_query_sheet(ws6, conn, """
        SELECT rq.review_id, p.filename, rq.source_table, rq.source_id,
               rq.issue, rq.context, rq.resolved, rq.resolution
        FROM review_queue rq
        JOIN papers p ON rq.paper_id = p.paper_id
        WHERE rq.resolved = 0
        ORDER BY rq.review_id
    """, headers=[
        'ID', 'File', 'Source Table', 'Source ID',
        'Issue', 'Context', 'Resolved', 'Resolution',
    ])

    # Save
    wb.save(output_path)
    conn.close()

    if verbose:
        print(f"Exported to: {output_path}")

    return output_path


def _write_headers(ws, headers):
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_query_sheet(ws, conn, query, headers):
    _write_headers(ws, headers)
    rows = conn.execute(query).fetchall()
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    _auto_width(ws)


def _auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:50]:  # check first 50 rows only
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_summary(ws, conn):
    ws.cell(row=1, column=1, value="MABE Paper Data Extraction Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    from paper_extractor.db import get_stats
    stats = get_stats(conn)

    rows = [
        ("Papers processed", stats.get('papers', 0)),
        ("Binding data entries", stats.get('binding_data', 0)),
        ("Physical constants", stats.get('physical_constants', 0)),
        ("Crystal contacts", stats.get('contacts', 0)),
        ("Raw tables stored", stats.get('raw_tables', 0)),
        ("Items needing review", stats.get('unresolved_reviews', 0)),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=1).font = Font(bold=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
'''

FILES["paper_extractor/extractor.py"] = r'''
"""
paper_extractor/extractor.py -- Main extraction pipeline
=========================================================

Orchestrates: PDF reading -> table classification -> data parsing -> DB storage.
"""

import os
import sys
from pathlib import Path
from typing import List, Dict, Any

from paper_extractor.db import (
    init_db, get_db_path, file_hash, is_paper_processed,
    register_paper, insert_binding, insert_physical_constant,
    insert_raw_table, add_review_item, get_stats,
)
from paper_extractor.pdf_reader import read_pdf, extract_metadata_from_text
from paper_extractor.patterns import (
    classify_table, parse_binding_table,
    extract_binding_from_text, extract_constants_from_text,
    parse_garbled_table_rows,
)


def process_folder(input_dir: str, db_dir: str = None, verbose: bool = True) -> Dict[str, int]:
    """
    Process all PDFs in input_dir. Store results in db_dir.

    Parameters
    ----------
    input_dir : str
        Folder containing PDF files.
    db_dir : str or None
        Where to put the database. Default: same as input_dir.
    verbose : bool
        Print progress.

    Returns
    -------
    dict with counts of new papers, tables, binding entries, etc.
    """
    if db_dir is None:
        db_dir = input_dir

    db_path = get_db_path(db_dir)
    conn = init_db(db_path)

    pdfs = sorted(Path(input_dir).glob("*.pdf"))
    if not pdfs:
        if verbose:
            print(f"No PDF files found in {input_dir}")
        return {'new_papers': 0}

    counts = {
        'total_pdfs': len(pdfs),
        'skipped': 0,
        'processed': 0,
        'failed': 0,
        'tables_found': 0,
        'tables_classified': 0,
        'binding_entries': 0,
        'constant_entries': 0,
        'review_items': 0,
    }

    for pdf_path in pdfs:
        fhash = file_hash(str(pdf_path))

        if is_paper_processed(conn, fhash):
            counts['skipped'] += 1
            if verbose:
                print(f"  SKIP (already processed): {pdf_path.name}")
            continue

        if verbose:
            print(f"  Processing: {pdf_path.name}")

        try:
            result = process_single_paper(str(pdf_path), conn, verbose=verbose)
            counts['processed'] += 1
            counts['tables_found'] += result.get('tables_found', 0)
            counts['tables_classified'] += result.get('tables_classified', 0)
            counts['binding_entries'] += result.get('binding_entries', 0)
            counts['constant_entries'] += result.get('constant_entries', 0)
            counts['review_items'] += result.get('review_items', 0)
        except Exception as e:
            counts['failed'] += 1
            if verbose:
                print(f"    FAILED: {e}")

    conn.close()

    if verbose:
        print(f"\nDone. {counts['processed']} new papers processed, "
              f"{counts['skipped']} skipped, {counts['failed']} failed.")
        print(f"  Tables found: {counts['tables_found']}")
        print(f"  Binding entries: {counts['binding_entries']}")
        print(f"  Constants: {counts['constant_entries']}")
        if counts['review_items'] > 0:
            print(f"  Items needing review: {counts['review_items']}")

    return counts


def process_single_paper(filepath: str, conn, verbose: bool = False) -> Dict[str, int]:
    """Process one PDF file."""
    result = {
        'tables_found': 0,
        'tables_classified': 0,
        'binding_entries': 0,
        'constant_entries': 0,
        'review_items': 0,
    }

    # Read PDF
    content = read_pdf(filepath)
    if verbose:
        print(f"    Pages: {content.n_pages}, Tables: {len(content.all_tables)}")

    # Extract metadata
    meta = extract_metadata_from_text(content.pages[0].text if content.pages else "")
    meta.update({k: v for k, v in content.metadata.items() if v})

    # Register paper
    fhash = file_hash(filepath)
    paper_id = register_paper(
        conn,
        filename=content.filename,
        fhash=fhash,
        n_pages=content.n_pages,
        title=meta.get('title', ''),
        doi=meta.get('doi', ''),
        year=int(meta['year']) if meta.get('year', '').isdigit() else None,
    )

    # Process tables
    for table in content.all_tables:
        result['tables_found'] += 1

        # Classify
        classification, confidence = classify_table(
            table.headers,
            table.rows[:3] if table.rows else None,
        )

        # Store raw table
        raw_id = insert_raw_table(
            conn, paper_id,
            page_number=table.page_number,
            table_index=table.table_index,
            headers=table.headers,
            rows=table.rows,
            classified_as=classification,
            confidence=f"{confidence:.2f}",
        )

        if classification == 'other':
            continue

        result['tables_classified'] += 1

        if verbose:
            print(f"    Table p{table.page_number}#{table.table_index}: "
                  f"{classification} (conf={confidence:.2f}, "
                  f"{table.n_cols}cols x {table.n_rows}rows)")

        # Parse based on classification
        if classification == 'binding':
            entries = parse_binding_table(
                table.headers, table.rows,
                default_receptor=meta.get('title', ''),
            )
            for entry in entries:
                entry['table_ref'] = f"p{table.page_number} table {table.table_index}"
                bid = insert_binding(conn, paper_id, entry)
                if bid:
                    result['binding_entries'] += 1

            if not entries and confidence > 0.3:
                # Table looked like binding data but couldn't parse
                add_review_item(
                    conn, paper_id,
                    source_table='raw_tables', source_id=raw_id,
                    issue='binding_table_unparsed',
                    context=f"Headers: {table.headers[:5]}, {table.n_rows} rows",
                )
                result['review_items'] += 1

    # Garbled table parsing: try to extract from raw table text
    # that wasn't successfully parsed by the structured parser
    for table in content.all_tables:
        # Build raw text from all rows
        all_cells = []
        for row in table.rows:
            all_cells.extend(str(c) for c in row if c)
        raw_text = ' | '.join(all_cells)

        if len(raw_text) > 20:
            garbled_entries = parse_garbled_table_rows(raw_text)
            for entry in garbled_entries:
                entry['table_ref'] = f"garbled p{table.page_number}#{table.table_index}"
                bid = insert_binding(conn, paper_id, entry)
                if bid:
                    result['binding_entries'] += 1

    # Text-based extraction (lower confidence)
    # Process full document text as one block for better context
    full_text = content.full_text
    text_bindings = extract_binding_from_text(full_text)
    for entry in text_bindings:
        entry['table_ref'] = 'text (full doc)'
        bid = insert_binding(conn, paper_id, entry)
        if bid:
            result['binding_entries'] += 1

    text_constants = extract_constants_from_text(full_text)
    for entry in text_constants:
        entry['table_ref'] = 'text (full doc)'
        cid = insert_physical_constant(conn, paper_id, entry)
        if cid:
            result['constant_entries'] += 1

    return result
'''

FILES["paper_extractor/patterns.py"] = r'''
"""
paper_extractor/patterns.py -- Table classification and value extraction
=========================================================================

Rules-based pattern matching for identifying and parsing data from
extracted PDF tables and text. Each data type has:
  - Header patterns for table classification
  - Value extraction regex for parsing cells
  - Unit normalization
"""

import re
from typing import List, Dict, Any, Optional, Tuple


# =====================================================================
# TABLE CLASSIFICATION
# =====================================================================
# Score each table against data type patterns. Highest score wins.

# Header keywords that indicate binding thermodynamics
BINDING_KEYWORDS = [
    r'K[_\s]?[aAdD]', r'[Kk]_?(?:ass|dis)', r'association',
    r'dissociation', r'[\-−]?\s*[Δδ∆][GgHh]', r'delta\s*[GgHh]',
    r'[Tt][Δδ∆][Ss]',
    r'enthalpy', r'entropy', r'free\s*energy',
    r'[Kk]\s*\(?\s*M\s*[\-−]', r'kcal', r'kJ',
    r'IC\s*50', r'EC\s*50', r'binding',
    r'stoichiometry', r'\bn\b.*site',
]

# Header keywords for physical constants / dissolution / solvation
PHYSICAL_KEYWORDS = [
    r'[Δδ∆].*[Hh].*sol', r'dissolution', r'solvation',
    r'hydration', r'transfer', r'sublimation',
    r'[Cc][Pp]', r'heat\s*capacity', r'barrier',
    r'torsion', r'rotation', r'frequency',
    r'refractive', r'density', r'viscosity',
    r'solubility', r'activity\s*coefficient',
]

# Header keywords for crystal contacts
CONTACT_KEYWORDS = [
    r'hydrogen\s*bond', r'[Hh]-bond', r'contact',
    r'distance.*[ÅA]', r'angle.*deg',
    r'residue', r'atom', r'interaction',
    r'[Cc][Hh].?[ππ]', r'stacking', r'van\s*der\s*[Ww]aals',
    r'coordination', r'water.*bridge',
]


def classify_table(headers: List[str], first_rows: List[List[str]] = None) -> Tuple[str, float]:
    """
    Classify a table by its headers (and optionally first rows).

    Returns (classification, confidence) where classification is one of:
      'binding', 'physical', 'contacts', 'other'
    and confidence is 0-1.
    """
    header_text = ' '.join(h.lower() for h in headers if h)

    # Also check first data row for clues
    row_text = ''
    if first_rows and first_rows[0]:
        row_text = ' '.join(str(c).lower() for c in first_rows[0] if c)

    combined = header_text + ' ' + row_text

    scores = {
        'binding': _keyword_score(combined, BINDING_KEYWORDS),
        'physical': _keyword_score(combined, PHYSICAL_KEYWORDS),
        'contacts': _keyword_score(combined, CONTACT_KEYWORDS),
    }

    # Disambiguation: if both binding and physical score, check for
    # distinguishing terms
    if scores['binding'] > 0 and scores['physical'] > 0:
        # "sol", "dissolution", "hydration", "transfer" strongly favor physical
        phys_strong = len(re.findall(
            r'_sol\b|sol[uv]|dissolut|hydrat|transfer|sublim|barrier|torsion|rotat|heat.?cap',
            combined, re.IGNORECASE))
        # "Ka", "Kd", "IC50", "binding", "association" strongly favor binding
        bind_strong = len(re.findall(
            r'K[_\s]?[aAdD]|IC\s*50|associat|dissociat|binding\s*constant',
            combined, re.IGNORECASE))
        scores['physical'] += phys_strong * 2
        scores['binding'] += bind_strong * 2

    best = max(scores, key=scores.get)
    best_score = scores[best]

    if best_score < 0.5:
        return 'other', best_score

    # Normalize to 0-1
    confidence = min(1.0, best_score / 3.0)
    return best, confidence


def _keyword_score(text: str, patterns: List[str]) -> float:
    """Count keyword matches."""
    score = 0
    for pattern in patterns:
        if re.search(pattern, text, re.IGNORECASE):
            score += 1
    return score


# =====================================================================
# VALUE EXTRACTION
# =====================================================================

def parse_number(s: str) -> Optional[float]:
    """
    Parse a number from a table cell. Handles:
      - Plain numbers: 1.23, -4.56
      - Scientific notation: 1.2e4, 1.2 x 10^4, 1.2 × 10^4
      - Parenthetical uncertainty: 1.23(5), 1.23 ± 0.05
      - Unicode minus: −4.56
      - Comma thousands: 1,234
    Returns None if unparseable.
    """
    if not s or not s.strip():
        return None

    s = s.strip()

    # Non-binding indicators
    if s.upper() in ('NB', 'N.B.', 'ND', 'N.D.', '-', '—', '–', 'NI', 'N.I.'):
        return None

    # Unicode minus
    s = s.replace('−', '-').replace('–', '-').replace('—', '-')

    # Remove ± and everything after
    s = re.sub(r'\s*[±]\s*[\d.]+', '', s)

    # Remove parenthetical uncertainty: 1.23(5)
    s = re.sub(r'\(\d+\)', '', s)

    # Remove commas in numbers
    s = s.replace(',', '')

    # Scientific notation: "1.2 x 10^4" or "1.2 × 10^4" or "1.2 × 10 4"
    sci_match = re.match(r'([-]?\d+\.?\d*)\s*[×xX]\s*10\s*[\^]?\s*([-]?\d+)', s)
    if sci_match:
        mantissa = float(sci_match.group(1))
        exponent = int(sci_match.group(2))
        return mantissa * (10 ** exponent)

    # Plain scientific: 1.2e4
    try:
        return float(s)
    except ValueError:
        pass

    # Try extracting first number
    num_match = re.search(r'([-]?\d+\.?\d*(?:[eE][+-]?\d+)?)', s)
    if num_match:
        try:
            return float(num_match.group(1))
        except ValueError:
            pass

    return None


def detect_unit(header: str, cell: str = "") -> Optional[str]:
    """
    Detect unit from a column header or cell value.
    """
    combined = (header + ' ' + cell).lower()

    # Energy units
    if re.search(r'kj\s*/?\s*mol|kj\s*mol', combined):
        return 'kJ/mol'
    if re.search(r'kcal\s*/?\s*mol|kcal\s*mol', combined):
        return 'kcal/mol'
    if re.search(r'cal\s*/?\s*mol\s*/?\s*k|cal\s*mol.*k', combined):
        return 'cal/mol/K'

    # Concentration
    if re.search(r'\bm[−\-]1\b|m\s*[\-−]\s*1|\bm\^?\s*[\-−]\s*1', combined):
        return 'M^-1'
    if re.search(r'\bmm\b|millimol', combined):
        return 'mM'
    if re.search(r'[μµ]m\b|micromol', combined):
        return 'uM'

    # Distance
    if re.search(r'[Åå]|angstrom', combined):
        return 'A'
    if re.search(r'\bnm\b', combined):
        return 'nm'

    # Angle
    if re.search(r'deg|°', combined):
        return 'deg'

    # Temperature
    if re.search(r'[°]?\s*c\b|celsius', combined):
        return 'C'
    if re.search(r'\bk\b|kelvin', combined):
        return 'K'

    return None


def normalize_energy_to_kj(value: float, unit: str) -> float:
    """Convert energy to kJ/mol."""
    if unit == 'kcal/mol':
        return value * 4.184
    if unit == 'kJ/mol':
        return value
    if unit == 'cal/mol':
        return value * 4.184e-3
    return value  # unknown unit, return as-is


# =====================================================================
# BINDING TABLE PARSER
# =====================================================================

# Column header mapping: what header text maps to which field
BINDING_HEADER_MAP = {
    'Ka': [r'K\s*a', r'K\s*ass', r'association\s*constant'],
    'Kd': [r'K\s*d', r'K\s*dis', r'dissociation\s*constant', r'IC\s*50'],
    'dG': [r'[Δδ∆d]\s*G', r'free\s*energy', r'\-?\s*[Δδ∆d]G'],
    'dH': [r'[Δδ∆d]\s*H', r'enthalpy', r'\-?\s*[Δδ∆d]H'],
    'TdS': [r'T\s*[Δδ∆d]\s*S', r'[Δδ∆d]S', r'entropy', r'\-?\s*T[Δδ∆d]S'],
    'dCp': [r'[Δδ∆d]\s*C\s*p', r'heat\s*capacity'],
    'n_sites': [r'\bn\b', r'stoich', r'sites?\s*/?\s*mono'],
    'ligand': [r'ligand', r'sugar', r'saccharide', r'carbohydrate', r'compound', r'substrate'],
    'receptor': [r'lectin', r'protein', r'receptor', r'host'],
}


def map_headers_to_fields(headers: List[str]) -> Dict[int, str]:
    """
    Map column indices to field names.
    Returns {col_index: field_name}.
    """
    mapping = {}
    for i, header in enumerate(headers):
        if not header:
            continue
        h_lower = header.lower().strip()
        for field, patterns in BINDING_HEADER_MAP.items():
            for pattern in patterns:
                if re.search(pattern, h_lower, re.IGNORECASE):
                    if field not in mapping.values():  # don't double-assign
                        mapping[i] = field
                    break
    return mapping


def parse_binding_table(headers: List[str], rows: List[List[str]],
                        default_receptor: str = "",
                        default_method: str = "ITC") -> List[Dict[str, Any]]:
    """
    Parse a table classified as binding data.
    Returns list of dicts ready for db.insert_binding().
    """
    col_map = map_headers_to_fields(headers)
    if not col_map:
        return []

    # Detect units from headers
    units = {}
    for col_idx, field in col_map.items():
        if field in ('Ka', 'Kd', 'dG', 'dH', 'TdS', 'dCp'):
            unit = detect_unit(headers[col_idx])
            if unit:
                units[field] = unit

    results = []
    for row in rows:
        if not row or all(not c.strip() for c in row if c):
            continue

        entry = {
            'receptor': default_receptor,
            'method': default_method,
            'confidence': 'medium',
        }

        for col_idx, field in col_map.items():
            if col_idx >= len(row):
                continue
            cell = row[col_idx]

            if field in ('ligand', 'receptor'):
                entry[field] = cell.strip()
            elif field in ('Ka', 'Kd', 'dG', 'dH', 'TdS', 'dCp', 'n_sites'):
                val = parse_number(cell)
                if val is not None:
                    entry[field] = val
                    if field in units:
                        entry[f'{field}_unit'] = units[field]

        # Only keep if we got at least ligand + one numeric value
        has_numeric = any(isinstance(entry.get(f), (int, float))
                         for f in ('Ka', 'Kd', 'dG', 'dH', 'TdS'))
        has_ligand = bool(entry.get('ligand', '').strip())

        if has_numeric and has_ligand:
            entry['raw_text'] = ' | '.join(str(c) for c in row if c)
            results.append(entry)

    return results


# =====================================================================
# TEXT PREPROCESSING
# =====================================================================

def normalize_text(text: str) -> str:
    """
    Normalize PDF-extracted text for pattern matching.
    Fixes common PDF artifacts that break regex.
    """
    # Replace newlines between parts of the same token
    # e.g. "K\na\n=" -> "Ka ="
    t = text

    # Fix split Ka/Kd/dH/dG tokens
    t = re.sub(r'K\s*\n\s*a\b', 'Ka', t)
    t = re.sub(r'K\s*\n\s*d\b', 'Kd', t)
    t = re.sub(r'[Δδ∆]\s*\n\s*H\b', 'ΔH', t)
    t = re.sub(r'[Δδ∆]\s*\n\s*G\b', 'ΔG', t)
    t = re.sub(r'[Δδ∆]\s*\n\s*S\b', 'ΔS', t)
    t = re.sub(r'T\s*[Δδ∆]\s*\n?\s*S\b', 'TΔS', t)

    # Collapse runs of whitespace (but preserve paragraph breaks)
    t = re.sub(r'[ \t]+', ' ', t)
    # Collapse single newlines (keep double = paragraph break)
    t = re.sub(r'(?<!\n)\n(?!\n)', ' ', t)

    # Fix "minus" artifacts
    t = t.replace('−', '-').replace('–', '-')

    # Fix "x10" scientific notation: "4.9 × 10^5" -> "4.9e5"
    t = re.sub(r'(\d)\s*[×xX]\s*10\s*[\^]?\s*([\-]?\d+)', r'\1e\2', t)

    return t


def get_context(text: str, pos: int, window: int = 150) -> str:
    """Get surrounding text for context, respecting sentence boundaries."""
    start = max(0, pos - window)
    end = min(len(text), pos + window)
    chunk = text[start:end].replace('\n', ' ').strip()

    # Try to trim to sentence containing the match
    # Find the sentence boundary before and after the match position
    relative_pos = pos - start
    # Look backward for sentence start
    sent_start = 0
    for delim_pos in range(relative_pos, 0, -1):
        if chunk[delim_pos] in '.!?' and delim_pos < relative_pos - 2:
            sent_start = delim_pos + 1
            break
    # Look forward for sentence end
    sent_end = len(chunk)
    for delim_pos in range(relative_pos, len(chunk)):
        if chunk[delim_pos] in '.!?' and delim_pos > relative_pos + 5:
            sent_end = delim_pos + 1
            break

    return chunk[sent_start:sent_end].strip()


# =====================================================================
# TEXT-BASED EXTRACTION (for data mentioned in prose, not tables)
# =====================================================================

def extract_binding_from_text(text: str) -> List[Dict[str, Any]]:
    """
    Extract binding data mentioned in running text.
    Handles garbled PDF whitespace, split tokens, missing spaces.
    """
    results = []
    seen_values = set()  # dedup by (field, value)

    # Normalize text first
    t = normalize_text(text)

    # ── PATTERN SET 1: Ka/Kd = VALUE ──────────────────────────────────
    # Handles: "Ka = 18,600 M-1", "Ka=3981M-1", "Ka of 1.2 x 10^4"
    # Also: "Ka (M-1)" followed by value on next line
    ka_patterns = [
        # "Ka = VALUE M-1" or "Ka=VALUE" (with optional spaces)
        re.compile(
            r'Ka\s*(?:=|is|was|of|:)\s*'
            r'([\d.,]+(?:[eE][\-]?\d+|\s*[×xX]\s*10\s*[\^]?\s*[\-]?\d+)?)\s*'
            r'(M\s*[\-]\s*1|M-1)?',
            re.IGNORECASE
        ),
        # "Kd = VALUE mM/uM/nM"
        re.compile(
            r'Kd\s*(?:=|is|was|of|:)\s*'
            r'([\d.,]+(?:[eE][\-]?\d+|\s*[×xX]\s*10\s*[\^]?\s*[\-]?\d+)?)\s*'
            r'(mM|[μµu]M|nM|M)',
            re.IGNORECASE
        ),
        # "association constant" or "binding constant" followed by value
        re.compile(
            r'(?:association|binding)\s*constants?\s*(?:\([^)]*\))?\s*'
            r'(?:=|is|was|of|:)\s*'
            r'([\d.,]+(?:[eE][\-]?\d+|\s*[×xX]\s*10\s*[\^]?\s*[\-]?\d+)?)\s*'
            r'(M\s*[\-]\s*1)?',
            re.IGNORECASE
        ),
        # "Ka VALUE" with number following (no = sign, common in garbled text)
        re.compile(
            r'Ka\s+'
            r'([\d.,]+(?:[eE][\-]?\d+|\s*[×xX]\s*10\s*[\^]?\s*[\-]?\d+)?)\s*'
            r'(M\s*[\-]\s*1)?',
        ),
    ]

    for pattern in ka_patterns:
        for match in pattern.finditer(t):
            val = parse_number(match.group(1))
            if val is None or val == 0:
                continue
            # Skip if too small (likely not Ka) or already seen
            key = ('Ka', round(val, 1))
            if key in seen_values:
                continue
            seen_values.add(key)

            context = get_context(t, match.start())
            ligand, receptor = _extract_context_names(context)

            results.append({
                'receptor': receptor,
                'ligand': ligand,
                'Ka': val,
                'method': 'text_extraction',
                'confidence': 'low',
                'raw_text': match.group(0)[:200].strip(),
                'conditions': context[:200],
            })

    # ── PATTERN SET 2: ΔH = VALUE kcal/mol ────────────────────────────
    dh_patterns = [
        # "ΔH = -8.4 kcal/mol" or "ΔHwas-6.6kcal/mol" (no spaces)
        re.compile(
            r'[Δδ∆]H\s*(?:°\s*)?(?:=|is|was|of|:)\s*'
            r'([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol)',
            re.IGNORECASE
        ),
        # "enthalpy of -8.4 kcal/mol"
        re.compile(
            r'enthalpy\s+(?:of\s+)?(?:binding\s+)?'
            r'(?:=|is|was|of|:)?\s*'
            r'([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol)',
            re.IGNORECASE
        ),
        # "-ΔH = 8.4" or "-ΔH (kcal/mol)" patterns
        re.compile(
            r'[\-]?\s*[Δδ∆]H\s*(?:\([^)]*\))?\s*(?:=|:)\s*'
            r'([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol)?',
            re.IGNORECASE
        ),
    ]

    for pattern in dh_patterns:
        for match in pattern.finditer(t):
            val = parse_number(match.group(1))
            if val is None:
                continue
            unit = match.group(2) if match.lastindex >= 2 and match.group(2) else 'kcal/mol'
            key = ('dH', round(val, 2))
            if key in seen_values:
                continue
            seen_values.add(key)

            context = get_context(t, match.start())
            ligand, receptor = _extract_context_names(context)

            results.append({
                'receptor': receptor,
                'ligand': ligand,
                'dH': val,
                'dH_unit': unit.replace('/', '/'),
                'method': 'text_extraction',
                'confidence': 'low',
                'raw_text': match.group(0)[:200].strip(),
                'conditions': context[:200],
            })

    # ── PATTERN SET 3: ΔG = VALUE ─────────────────────────────────────
    dg_patterns = [
        re.compile(
            r'[Δδ∆]G\s*(?:°\s*)?(?:=|is|was|of|:)\s*'
            r'([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol)',
            re.IGNORECASE
        ),
        re.compile(
            r'free\s*energy\s+(?:of\s+)?(?:binding\s+)?'
            r'(?:=|is|was|of|:)?\s*'
            r'([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol)',
            re.IGNORECASE
        ),
    ]

    for pattern in dg_patterns:
        for match in pattern.finditer(t):
            val = parse_number(match.group(1))
            if val is None:
                continue
            unit = match.group(2) if match.lastindex >= 2 and match.group(2) else 'kcal/mol'
            key = ('dG', round(val, 2))
            if key in seen_values:
                continue
            seen_values.add(key)

            context = get_context(t, match.start())
            ligand, receptor = _extract_context_names(context)

            results.append({
                'receptor': receptor,
                'ligand': ligand,
                'dG': val,
                'dG_unit': unit,
                'method': 'text_extraction',
                'confidence': 'low',
                'raw_text': match.group(0)[:200].strip(),
                'conditions': context[:200],
            })

    # ── PATTERN SET 4: TΔS = VALUE ────────────────────────────────────
    tds_patterns = [
        re.compile(
            r'T\s*[Δδ∆]S\s*(?:°\s*)?(?:=|is|was|of|:)\s*'
            r'([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol)',
            re.IGNORECASE
        ),
    ]

    for pattern in tds_patterns:
        for match in pattern.finditer(t):
            val = parse_number(match.group(1))
            if val is None:
                continue
            unit = match.group(2) if match.lastindex >= 2 and match.group(2) else 'kcal/mol'
            key = ('TdS', round(val, 2))
            if key in seen_values:
                continue
            seen_values.add(key)

            context = get_context(t, match.start())
            ligand, receptor = _extract_context_names(context)

            results.append({
                'receptor': receptor,
                'ligand': ligand,
                'TdS': val,
                'TdS_unit': unit,
                'method': 'text_extraction',
                'confidence': 'low',
                'raw_text': match.group(0)[:200].strip(),
                'conditions': context[:200],
            })

    return results


# =====================================================================
# CONTEXT NAME EXTRACTION
# =====================================================================

# Known sugar names for context matching
KNOWN_SUGARS = [
    'glucose', 'mannose', 'galactose', 'fructose', 'fucose', 'xylose',
    'ribose', 'lactose', 'sucrose', 'maltose', 'cellobiose', 'chitobiose',
    'GlcNAc', 'GalNAc', 'LacNAc', 'ManNAc', 'Neu5Ac', 'sialic',
    'MeαMan', 'MeαGlc', 'MeαGal', 'MeβGal', 'MeβGlc',
    'MeRMan', 'MeRGlc', 'MeRGal', 'MebGal', 'MebGlc',
    'trimannoside', 'mannotriose', 'mannobiose', 'chitotriose',
    'D-glucose', 'D-mannose', 'D-galactose', 'D-fructose', 'D-xylose',
    'D-ribose', 'D-cellobiose',
    'methyl α-D-mannopyranoside', 'methyl α-D-glucopyranoside',
    '2-deoxy', '3-deoxy', '4-deoxy', '6-deoxy',
    'thiodigalactoside', 'dithiogalactoside',
    'T-antigen', 'blood group',
]

KNOWN_RECEPTORS = [
    'ConA', 'concanavalin', 'WGA', 'wheat germ',
    'DGL', 'Dioclea', 'galectin', 'PNA', 'peanut',
    'SBA', 'soybean', 'ECorL', 'Erythrina',
    'hevein', 'UDA', 'Urtica', 'GNA', 'Galanthus',
    'ASA', 'Allium', 'NPL', 'Narcissus',
    'artocarpin', 'banana lectin', 'RCA', 'Ricinus',
    'abrin', 'jacalin', 'MBP', 'mannose binding',
    'lysozyme', 'receptor', 'lectin',
    'TC14', 'selectin',
]


def _extract_context_names(context: str) -> tuple:
    """
    Extract ligand and receptor names from surrounding text context.
    Returns (ligand, receptor).
    """
    ligand = ''
    receptor = ''

    ctx_lower = context.lower()

    # Find sugar/ligand
    for sugar in KNOWN_SUGARS:
        if sugar.lower() in ctx_lower:
            ligand = sugar
            break

    # Find receptor
    for rec in KNOWN_RECEPTORS:
        if rec.lower() in ctx_lower:
            receptor = rec
            break

    return ligand, receptor


# =====================================================================
# GARBLED TABLE ROW PARSER
# =====================================================================
# Handles tables where pdfplumber jams rows together, like:
# "D-Glucose10 1 ... | 8,000 500 ... | 18,600 7900 5300 5800 725 180 140 220 60 30"

def parse_garbled_table_rows(raw_text: str, headers_hint: str = "") -> List[Dict[str, Any]]:
    """
    Parse garbled table text where structure is destroyed but data is present.

    Handles patterns like:
      "D-Glucose10 1 D-Galactose15 1 | 8,000 30 | 18,600 180"
    where column 1 has names+numbers jammed together, and other columns
    have the Ka values separated by spaces.
    """
    results = []

    # Split on pipe characters
    columns = [c.strip() for c in raw_text.split('|') if c.strip()]
    if len(columns) < 2:
        return results

    # ── Step 1: Find the column with sugar names ──
    name_col = None
    num_cols = []

    for i, col in enumerate(columns):
        has_sugar = any(s.lower() in col.lower() for s in
                       ['glucose', 'mannose', 'galactose', 'fructose', 'fucose',
                        'xylose', 'ribose', 'lactose', 'cellobiose', 'GlcNAc',
                        'GalNAc', 'LacNAc', 'manno', 'deoxy', 'Glc', 'Man',
                        'Gal', 'Fuc', 'Xyl'])
        if has_sugar:
            name_col = i
        else:
            # Check if this column is mostly numbers
            nums = re.findall(r'[\d,]+(?:\.\d+)?', col)
            if nums:
                num_cols.append(i)

    if name_col is None or not num_cols:
        return results

    # ── Step 2: Split name column into individual entries ──
    name_text = columns[name_col]

    # Pattern: "D-Glucose10" or "2-Deoxy-D-Glucose14" — name followed by
    # a compound number, then maybe a space and another small number
    # Split at transitions from letter/hyphen to digit-digit pattern
    # that marks a new compound number
    sugar_entries = re.split(
        r'(?<=\d)\s+(?=[A-Z2-9])',  # split at "...10 1 D-..." boundaries
        name_text
    )

    # Better: find all "Name + CompoundNumber" patterns
    sugar_names_list = [
        'Glucose', 'Mannose', 'Galactose', 'Fructose', 'Fucose',
        'Xylose', 'Ribose', 'Cellobiose', 'Lactose', 'Maltose',
        'Sucrose', 'Chitobiose', 'GlcNAc', 'GalNAc', 'LacNAc',
        'Glucoside', 'Mannoside', 'Galactoside', 'GlucuronicAcid',
    ]
    sugar_alts = '|'.join(sugar_names_list)
    name_pattern = re.compile(
        r'((?:Methyl|2-Deoxy|3-Deoxy|4-Deoxy|6-Deoxy|N-Acetyl|[DL]-|[αβß]-)*'
        r'(?:' + sugar_alts + r'))'
        r'\s*(\d{0,3})',
        re.IGNORECASE
    )

    sugar_names = []
    for match in name_pattern.finditer(name_text):
        name = match.group(1).strip()
        # Clean up concatenated names
        name = re.sub(r'([a-z])([A-Z])', r'\1 \2', name)  # camelCase split
        sugar_names.append(name)

    if not sugar_names:
        return results

    # ── Step 3: Extract numbers from the last numeric column (usually ITC) ──
    # Prefer the last numeric column (often ITC Ka values)
    best_num_col = num_cols[-1]
    num_text = columns[best_num_col]
    numbers = re.findall(r'[\d,]+(?:\.\d+)?', num_text)
    parsed_numbers = [parse_number(n) for n in numbers]
    parsed_numbers = [n for n in parsed_numbers if n is not None]

    # ── Step 4: Zip names with numbers ──
    n_pairs = min(len(sugar_names), len(parsed_numbers))
    for i in range(n_pairs):
        val = parsed_numbers[i]
        if val <= 0:
            continue
        results.append({
            'ligand': sugar_names[i],
            'Ka': val,
            'confidence': 'low',
            'method': 'garbled_table',
            'raw_text': f"{sugar_names[i]} -> {val} (col {best_num_col})",
        })

    return results


# =====================================================================
# PHYSICAL CONSTANT EXTRACTION FROM TEXT
# =====================================================================

def extract_constants_from_text(text: str) -> List[Dict[str, Any]]:
    """
    Extract physical constants mentioned in text.
    Handles garbled whitespace and various formats.
    """
    results = []
    seen = set()
    t = normalize_text(text)

    patterns = [
        # "ΔH = VALUE kcal/mol" or "ΔHwas-6.6kcal/mol"
        (re.compile(
            r'[Δδ∆]H\s*(?:°\s*)?(?:sol|hyd|sub|vap|fus|mix)?\s*'
            r'(?:=|is|was|of|:)\s*([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol|cal/?mol)',
            re.IGNORECASE),
         lambda m: {'quantity': 'dH', 'compound': '',
                     'value': parse_number(m.group(1)), 'unit': m.group(2)}),

        # "ΔG = VALUE"
        (re.compile(
            r'[Δδ∆]G\s*(?:°\s*)?(?:sol|hyd|sub|vap|fus|mix)?\s*'
            r'(?:=|is|was|of|:)\s*([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol|cal/?mol)',
            re.IGNORECASE),
         lambda m: {'quantity': 'dG', 'compound': '',
                     'value': parse_number(m.group(1)), 'unit': m.group(2)}),

        # "ΔCp = VALUE cal/mol/K"
        (re.compile(
            r'[Δδ∆]\s*C\s*p\s*(?:°\s*)?'
            r'(?:=|is|was|of|:)\s*([\-]?[\d.,]+)\s*'
            r'(cal/?mol/?K|J/?mol/?K|kJ/?mol/?K|kcal/?mol/?K)',
            re.IGNORECASE),
         lambda m: {'quantity': 'dCp', 'compound': '',
                     'value': parse_number(m.group(1)), 'unit': m.group(2)}),

        # "enthalpy of solution" type patterns
        (re.compile(
            r'(?:enthalpy|heat)\s+of\s+(?:solution|dissolution|hydration|solvation|transfer)\s+'
            r'(?:of\s+(\S+(?:\s+\S+){0,2}?)\s+)?'
            r'(?:=|is|was|:)\s*([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol|cal/?mol)',
            re.IGNORECASE),
         lambda m: {'quantity': 'dH_sol', 'compound': m.group(1) or '',
                     'value': parse_number(m.group(2)), 'unit': m.group(3)}),

        # Barrier height patterns
        (re.compile(
            r'(?:barrier|activation\s*energy)\s*(?:=|is|was|of|:)\s*'
            r'([\-]?[\d.,]+)\s*'
            r'(kcal/?mol|kJ/?mol)',
            re.IGNORECASE),
         lambda m: {'quantity': 'barrier', 'compound': '',
                     'value': parse_number(m.group(1)), 'unit': m.group(2)}),

        # pKa patterns
        (re.compile(
            r'pK\s*a\s*(?:=|is|was|of|:)\s*([\d.,]+)',
            re.IGNORECASE),
         lambda m: {'quantity': 'pKa', 'compound': '',
                     'value': parse_number(m.group(1)), 'unit': ''}),
    ]

    for pattern, extractor in patterns:
        for match in pattern.finditer(t):
            data = extractor(match)
            if data.get('value') is None:
                continue
            key = (data['quantity'], round(data['value'], 3))
            if key in seen:
                continue
            seen.add(key)

            context = get_context(t, match.start())
            data['confidence'] = 'low'
            data['raw_text'] = match.group(0)[:200].strip()
            # Try to get compound name from context
            if not data.get('compound'):
                for sugar in KNOWN_SUGARS:
                    if sugar.lower() in context.lower():
                        data['compound'] = sugar
                        break
            results.append(data)

    return results
'''

FILES["paper_extractor/pdf_reader.py"] = r'''
"""
paper_extractor/pdf_reader.py -- PDF text and table extraction
===============================================================
Uses pdfplumber for both text and table extraction.
Falls back to pypdf if pdfplumber fails on a page.
"""

import re
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None


@dataclass
class ExtractedTable:
    """A table extracted from a PDF page."""
    page_number: int
    table_index: int
    headers: List[str]
    rows: List[List[str]]
    raw_bbox: Optional[tuple] = None

    @property
    def n_cols(self) -> int:
        return len(self.headers) if self.headers else 0

    @property
    def n_rows(self) -> int:
        return len(self.rows)


@dataclass
class PageContent:
    """Content extracted from one PDF page."""
    page_number: int
    text: str
    tables: List[ExtractedTable] = field(default_factory=list)


@dataclass
class PaperContent:
    """All content from a PDF."""
    filename: str
    n_pages: int
    pages: List[PageContent] = field(default_factory=list)
    metadata: Dict[str, str] = field(default_factory=dict)

    @property
    def full_text(self) -> str:
        return "\n\n".join(p.text for p in self.pages)

    @property
    def all_tables(self) -> List[ExtractedTable]:
        tables = []
        for p in self.pages:
            tables.extend(p.tables)
        return tables


def read_pdf(filepath: str) -> PaperContent:
    """
    Extract text and tables from a PDF.
    Uses pdfplumber (preferred) with pypdf fallback.
    """
    if pdfplumber is None and PdfReader is None:
        raise ImportError("Need pdfplumber or pypdf. Install: pip install pdfplumber")

    if pdfplumber is not None:
        return _read_with_pdfplumber(filepath)
    else:
        return _read_with_pypdf(filepath)


def _read_with_pdfplumber(filepath: str) -> PaperContent:
    """Extract using pdfplumber (handles tables well)."""
    import os
    pages = []

    with pdfplumber.open(filepath) as pdf:
        n_pages = len(pdf.pages)
        metadata = pdf.metadata or {}

        for i, page in enumerate(pdf.pages):
            page_num = i + 1
            # Text
            text = page.extract_text() or ""

            # Tables
            extracted_tables = []
            try:
                raw_tables = page.extract_tables()
                for t_idx, raw_table in enumerate(raw_tables or []):
                    if not raw_table or len(raw_table) < 2:
                        continue
                    # Clean cells
                    cleaned = []
                    for row in raw_table:
                        cleaned.append([_clean_cell(c) for c in row])

                    # First row with content = headers
                    headers = cleaned[0]
                    data_rows = cleaned[1:]

                    # Skip tables that are mostly empty
                    total_cells = sum(len(r) for r in data_rows)
                    non_empty = sum(1 for r in data_rows for c in r if c.strip())
                    if total_cells > 0 and non_empty / total_cells < 0.2:
                        continue

                    extracted_tables.append(ExtractedTable(
                        page_number=page_num,
                        table_index=t_idx,
                        headers=headers,
                        rows=data_rows,
                    ))
            except Exception:
                pass  # table extraction can fail on complex layouts

            pages.append(PageContent(
                page_number=page_num,
                text=text,
                tables=extracted_tables,
            ))

    return PaperContent(
        filename=os.path.basename(filepath),
        n_pages=n_pages,
        pages=pages,
        metadata={str(k): str(v) for k, v in (metadata or {}).items()},
    )


def _read_with_pypdf(filepath: str) -> PaperContent:
    """Fallback: pypdf for text only (no table extraction)."""
    import os
    reader = PdfReader(filepath)
    pages = []

    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        pages.append(PageContent(page_number=i+1, text=text))

    return PaperContent(
        filename=os.path.basename(filepath),
        n_pages=len(reader.pages),
        pages=pages,
        metadata={str(k): str(v) for k, v in (reader.metadata or {}).items()},
    )


def _clean_cell(cell) -> str:
    """Clean a table cell value."""
    if cell is None:
        return ""
    s = str(cell).strip()
    # Normalize whitespace
    s = re.sub(r'\s+', ' ', s)
    # Common PDF artifacts
    s = s.replace('\x00', '').replace('\ufb01', 'fi').replace('\ufb02', 'fl')
    return s


def extract_metadata_from_text(text: str) -> Dict[str, str]:
    """
    Try to extract title, authors, DOI from first page text.
    Heuristic — not reliable on all formats.
    """
    meta = {}

    # DOI
    doi_match = re.search(r'(?:doi|DOI)[:\s]*([10]\.\d{4,}/[^\s]+)', text)
    if doi_match:
        meta['doi'] = doi_match.group(1).rstrip('.')

    # Year from common patterns
    year_match = re.search(r'(?:19|20)\d{2}', text[:2000])
    if year_match:
        meta['year'] = year_match.group()

    # Title: often the first substantial line (heuristic)
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    for line in lines[:10]:
        if len(line) > 20 and not re.match(r'^(?:doi|DOI|http|www|\d)', line):
            if not any(w in line.lower() for w in ['copyright', 'received', 'accepted',
                                                     'published', 'journal', 'volume']):
                meta['title'] = line[:200]
                break

    return meta
'''

def deploy():
    created = []
    for relpath, content in FILES.items():
        fullpath = os.path.join(os.getcwd(), relpath)
        os.makedirs(os.path.dirname(fullpath), exist_ok=True)
        with open(fullpath, "w", encoding="utf-8") as fh:
            fh.write(content.lstrip("\n"))
        created.append(relpath)
        print("  Created: " + relpath)
    print(str(len(created)) + " files created.")
    print("")
    print("To reprocess papers with improved patterns:")
    print("  del papers\\paper_data.db")
    print("  python -m paper_extractor run papers")
    print("  python -m paper_extractor export papers")

if __name__ == "__main__":
    deploy()